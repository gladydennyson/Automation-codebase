# speech recognition

import speech_recognition as speech
from openpyxl import load_workbook

wb = load_workbook('sample.xlsx')

# grab the active worksheet
ws = wb.active


#a lot of variables used for loops and conditions
voice = speech.Recognizer()

while True:
    #with microphone input as the input
    with speech.Microphone() as source:

        print("Say Command.")

        #variable = the input from the microphone
        audio = voice.listen(source)
        voiceRecog = voice.recognize_google(audio)
        try:
            #print the word that was heard
            print("\nYou said: " + voiceRecog)
        except speech.UnknownValueError:
            #if couldnt recognise then print this msg
            print("Sorry, we couldn't make that out. Try again!")

        except speech.RequestError as e:
            print("Could not request results from Google Speech Recognition service")


        #list of commands  for comparing   
        words = voiceRecog.split(" ")
        print(words)
        if "apply" in voiceRecog:
            if i==1:
                ws[p3] = '=SUM('+p1+':'+p2+')'
            elif i==2:
                ws[p2] = '=ABS('+p1+')'

            elif i==3:
                ws[p2] = '=EXP('+p1+')'

            elif i==4:
                ws[p2] = '=sqrt('+p1+')'

            elif i==5:
                ws[p2] = '=log('+p1+')'

            elif i==6:
                ws[p2] = '=ln('+p1+')'

            elif i==7:
                ws[p2] = '=fact('+p1+')'

            elif i==8:
                ws[p2] = '=lcm('+p1+':'+p2+')'
                
            # Save the file
            wb.save("sample.xlsx")
            print("Say EXIT to terminate.")

        elif "exit" in voiceRecog:
            break

        elif words[0] == "add":
            i = 1
            p1 = words[1]
            p2 = words[2]
            p3 = words[4]
            print(p1 + p2 + p3)
            print("\nCommand Found!\n")


        elif words[0] == "absolute":
            i=2
            p1 = words[1]
            p2 = words[3]
            print("\nCommand Found!\n")

        elif words[0] == "exponential":
            i=3
            p1 = words[1]
            p2 = words[3]
            print("\nCommand Found!\n")

        elif words[0] == "square" and word[1] == "root":
            i=4
            p1 = words[2]
            p2 = words[4]
            print("\nCommand Found!\n")

        elif words[0] == "log":
            i=5
            p1 = words[1]
            p2 = words[3]
            print("\nCommand Found!\n")

        elif words[0] == "natural" and word[1] == "log":
            i=6
            p1 = words[2]
            p2 = words[4]
            print("\nCommand Found!\n")

        elif words[0] == "factorial":
            i=7
            p1 = words[1]
            p2 = words[3]
            print("\nCommand Found!\n")

        elif words[0] == "LCM":
            i=8
            p1 = words[1]
            p2 = words[2]
            p3 = words[4]
            print("\nCommand Found!\n")
            
        else:
            print("\nCommand not found!\n\n")

print("\n Done")


# Save the file
wb.save("sample.xlsx")

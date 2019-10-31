from openpyxl import load_workbook

wb = load_workbook('sample.xlsx')

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A2'] = 'Tom'
ws['B2'] = 3

ws['A3'] = 'Ma'
ws['B5'] = 2



# Save the file
wb.save("sample.xlsx")
